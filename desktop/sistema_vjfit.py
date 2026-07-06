"""
VJ FIT — Sistema de Vendas & Estoque  v6
─────────────────────────────────────────
NOVIDADE: Variantes de produto (Cor + Tamanho)
  • Cada variante tem código único (ex: 320001)
  • Agrupadas por Código Pai (ex: 32)
  • Baixa de estoque exata por cor/tamanho
  • Migração automática de backups antigos
LEITURA  → openpyxl   ESCRITA → xlwings
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import os, shutil, sys

# ── openpyxl ────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

# ── xlwings ──────────────────────────────────────────────────
try:
    import xlwings as xw
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "xlwings"])
    import xlwings as xw

# ── Pillow ───────────────────────────────────────────────────
try:
    from PIL import Image, ImageTk
    PIL_DISPONIVEL = True
except ImportError:
    PIL_DISPONIVEL = False

# ============================================================
# CONFIGURAÇÕES GLOBAIS
# ============================================================
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

ARQUIVO_EXCEL = os.path.join(BASE_DIR, "controle_pedidos.xlsx")
PASTA_FOTOS   = os.path.join(BASE_DIR, "fotos_produtos")
PASTA_BACKUP  = os.path.join(BASE_DIR, "backups")

FMT_MOEDA = 'R$ #.##0,00'
FMT_DATA  = 'DD/MM/AAAA'

# Cabeçalhos das abas (v6)
CAB_ESTOQUE = ["COD_PRODUTO","COD_PAI","NOME","CATEGORIA",
               "COR","TAMANHO","QUANTIDADE",
               "PRECO_CUSTO","PRECO_VENDA","FOTO_PATH","ULTIMA_ATUALIZACAO"]
CAB_VENDAS  = ["ID","COD_PRODUTO","NOME DO CLIENTE","PRODUTO",
               "COR","TAMANHO","QUANTIDADE",
               "VALOR UNIT.","VALOR TOTAL","DATA","PAGAMENTO","ENTREGA"]

# Índices (0-based) para leitura via openpyxl values_only
# ESTOQUE
I_E_COD=0; I_E_PAI=1; I_E_NOME=2; I_E_CAT=3; I_E_COR=4
I_E_TAM=5; I_E_QTD=6; I_E_CUSTO=7; I_E_VENDA=8; I_E_FOTO=9
# VENDAS (data na posição 9, pag=10, ent=11)
I_V_DATA=9


# ============================================================
# UTILITÁRIOS
# ============================================================
def para_float(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def normalizar_cod(cod):
    """
    Normaliza código removendo espaços e convertendo para maiúsculas.
    Para códigos puramente numéricos sem separadores: remove zeros à esquerda
    e casas decimais (ex: '01' → '1', '1.0' → '1').
    Para códigos com separadores ou letras: mantém como estão
    (ex: '1-01' → '1-01', 'CAM-P' → 'CAM-P', '10.01' → '10.01').
    Isso evita que '1', '10' e '100' se confundam entre si.
    """
    if cod is None: return ""
    s = str(cod).strip()
    # Só faz conversão inteira se for número puro sem separador
    # '1.0', '01', '100' → normaliza para int string
    # '1-01', '10.01', 'CAM-P' → mantém como está
    import re
    if re.fullmatch(r'\d+\.?\d*', s):
        try: return str(int(float(s)))
        except: pass
    return s.upper().strip()

def garantir_pastas():
    os.makedirs(PASTA_FOTOS,  exist_ok=True)
    os.makedirs(PASTA_BACKUP, exist_ok=True)


class ItemCarrinho:
    def __init__(self, cod, produto, cor, tam, qtd, valor):
        self.cod = cod
        self.produto = produto
        self.cor = cor
        self.tam = tam
        self.qtd = qtd
        self.valor = valor

    @property
    def total(self):
        return self.qtd * self.valor

class Carrinho:
    def __init__(self):
        self.itens = []

    def adicionar(self, item):
        # evita duplicado → soma quantidade
        for i in self.itens:
            if i.cod == item.cod:
                i.qtd += item.qtd
                return
        self.itens.append(item)

    def remover(self, index):
        if 0 <= index < len(self.itens):
            self.itens.pop(index)

    def limpar(self):
        self.itens.clear()

    def total(self):
        return sum(i.total for i in self.itens)

    def vazio(self):
        return len(self.itens) == 0


carrinho = Carrinho()

# ============================================================
# XLWINGS — ABERTURA INTELIGENTE
# ============================================================
def xw_abrir():
    nome_arquivo = os.path.basename(ARQUIVO_EXCEL).lower()
    try:
        for book in xw.books:
            try:
                if os.path.basename(book.fullname).lower() == nome_arquivo:
                    return book, False
            except Exception:
                pass
    except Exception:
        pass
    try:
        return xw.Book(ARQUIVO_EXCEL), True
    except Exception as e:
        raise Exception(
            f"Não foi possível abrir a planilha.\n"
            f"Verifique se '{os.path.basename(ARQUIVO_EXCEL)}' existe na mesma pasta do app.\n\nDetalhe: {e}"
        )

def xw_salvar_fechar(wb, deve_fechar):
    wb.save()
    if deve_fechar:
        wb.close()

def xw_ultima_linha(ws_xw, col="A"):
    return ws_xw.range(col + str(ws_xw.cells.last_cell.row)).end("up").row


# ============================================================
# MIGRAÇÃO DE BACKUP ANTIGO → v6
# Detecta formato antigo (8 colunas no ESTOQUE / 10 no VENDAS)
# e insere as colunas faltantes sem perder dados.
# ============================================================
def migrar_se_necessario():
    if not os.path.exists(ARQUIVO_EXCEL):
        return
    try:
        wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
        alterou = False

        # ── ESTOQUE ──────────────────────────────────────
        if "ESTOQUE" in wb.sheetnames:
            ws = wb["ESTOQUE"]
            headers = [str(ws.cell(1, c).value or "").strip()
                       for c in range(1, ws.max_column + 1)]

            if "COD_PAI" not in headers:
                # Formato antigo: COD_PRODUTO,NOME,CATEGORIA,QUANTIDADE,
                #                 PRECO_CUSTO,PRECO_VENDA,FOTO_PATH,ULTIMA_ATUALIZACAO
                # Novo:           COD_PRODUTO,COD_PAI,NOME,CATEGORIA,
                #                 COR,TAMANHO,QUANTIDADE,PRECO_CUSTO,PRECO_VENDA,FOTO_PATH,ULTIMA_ATUALIZACAO
                # Passos:
                #  1) Inserir coluna B (COD_PAI)
                #  2) Inserir coluna E (COR)
                #  3) Inserir coluna F (TAMANHO)  — após a COR já inserida

                ws.insert_cols(2)          # COD_PAI
                ws.cell(1, 2).value = "COD_PAI"
                ws.insert_cols(5)          # COR (antiga col D era QUANTIDADE → agora é col 5)
                ws.cell(1, 5).value = "COR"
                ws.insert_cols(6)          # TAMANHO
                ws.cell(1, 6).value = "TAMANHO"

                # Preenche COD_PAI com o mesmo valor de COD_PRODUTO
                for row in range(2, ws.max_row + 1):
                    cod = ws.cell(row, 1).value
                    if cod is not None:
                        ws.cell(row, 2).value = normalizar_cod(cod)

                _estilizar_cab_oxl(ws, 1, len(CAB_ESTOQUE), "1B5E20")
                alterou = True

        # ── VENDAS ───────────────────────────────────────
        if "VENDAS" in wb.sheetnames:
            ws = wb["VENDAS"]
            headers = [str(ws.cell(1, c).value or "").strip()
                       for c in range(1, ws.max_column + 1)]

            if "COR" not in headers:
                # Formato antigo: ID,COD_PRODUTO,NOME,PRODUTO,QUANTIDADE,VALOR UNIT.,VALOR TOTAL,DATA,PAG,ENTREGA
                # Novo:           ID,COD_PRODUTO,NOME,PRODUTO,COR,TAMANHO,QUANTIDADE,VALOR UNIT.,VALOR TOTAL,DATA,PAG,ENTREGA
                ws.insert_cols(5)          # COR
                ws.cell(1, 5).value = "COR"
                ws.insert_cols(6)          # TAMANHO
                ws.cell(1, 6).value = "TAMANHO"
                _estilizar_cab_oxl(ws, 1, len(CAB_VENDAS), "365F92")
                alterou = True

        if alterou:
            wb.save(ARQUIVO_EXCEL)
            wb.close()
            messagebox.showinfo(
                "✅ Planilha Atualizada",
                "Seu backup foi migrado automaticamente para a versão nova (v6).\n\n"
                "As colunas COR e TAMANHO foram adicionadas.\n"
                "Use 'Atualizar Estoque' para preencher cor/tamanho dos produtos existentes."
            )
        else:
            wb.close()
    except Exception as e:
        print(f"Erro na migração: {e}")


# ============================================================
# CRIAÇÃO INICIAL DA PLANILHA
# ============================================================
def _estilizar_cab_oxl(ws, linha, n_cols, cor):
    fill = PatternFill("solid", fgColor=cor)
    fonte = Font(bold=True, color="FFFFFF")
    al = Alignment(horizontal="center", vertical="center")
    for c in range(1, n_cols + 1):
        cell = ws.cell(linha, c)
        cell.fill, cell.font, cell.alignment = fill, fonte, al

def _ajustar_colunas(ws, extras=None):
    for col in ws.columns:
        mx = 0; cl = get_column_letter(col[0].column)
        for cell in col:
            try: mx = max(mx, len(str(cell.value or "")))
            except: pass
        ws.column_dimensions[cl].width = max(12, mx + 4)
    if extras:
        for cl, w in extras.items():
            ws.column_dimensions[cl].width = w

def _criar_tabela_oxl(ws, nome, ref, estilo="TableStyleMedium9"):
    for t in list(ws.tables.values()):
        if t.displayName == nome:
            del ws.tables[t.displayName]; break
    tab = Table(displayName=nome, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=estilo, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)

def _resumo_mensal_oxl(ws):
    fn_val  = Font(bold=True, size=11)
    fn_note = Font(italic=True, size=8, color="C62828")
    al_l = Alignment(horizontal="left",  vertical="center")
    al_r = Alignment(horizontal="right", vertical="center")
    al_c = Alignment(horizontal="center")
    # Resumo nas colunas N–P (dados em col J=DATA, I=VALOR TOTAL)
    ws["N1"] = "📊 RESUMO DO MÊS ATUAL"
    ws["N1"].font = Font(bold=True, size=12, color="1B5E20")
    ws["N1"].alignment = al_c
    try: ws.merge_cells("N1:P1")
    except: pass
    linhas = [
        ("N3","P3","💰 Total Vendido:",
         '=SUMPRODUCT((MONTH(J2:J5000)=MONTH(TODAY()))*(YEAR(J2:J5000)=YEAR(TODAY()))*(ISNUMBER(J2:J5000))*I2:I5000)',
         FMT_MOEDA),
        ("N4","P4","🛒 Nº de Vendas:",
         '=SUMPRODUCT((MONTH(J2:J5000)=MONTH(TODAY()))*(YEAR(J2:J5000)=YEAR(TODAY()))*(ISNUMBER(J2:J5000))*1)',
         "0"),
        ("N5","P5","🎯 Ticket Médio:",  '=IFERROR(P3/P4,0)', FMT_MOEDA),
        ("N6","P6","📦 Peças Vendidas:",
         '=SUMPRODUCT((MONTH(J2:J5000)=MONTH(TODAY()))*(YEAR(J2:J5000)=YEAR(TODAY()))*(ISNUMBER(J2:J5000))*G2:G5000)',
         "0"),
    ]
    for lab_c, val_c, lbl, formula, fmt in linhas:
        ws[lab_c] = lbl; ws[lab_c].font = fn_val; ws[lab_c].alignment = al_l
        ws[val_c] = formula; ws[val_c].number_format = fmt
        ws[val_c].font = fn_val; ws[val_c].alignment = al_r
    ws["N8"] = "Mês:"; ws["N8"].font = Font(italic=True, size=9, color="555555")
    ws["P8"] = '=TEXT(TODAY(),"MMMM/AAAA")'
    ws["P8"].font = Font(bold=True, size=10, color="1565C0"); ws["P8"].alignment = al_r
    ws["N9"] = "⚠ Feche o Excel para editar"; ws["N9"].font = fn_note
    try: ws.merge_cells("N9:P9")
    except: pass
    for cl, w in {"N": 26, "O": 4, "P": 18}.items():
        ws.column_dimensions[cl].width = w

def criar_planilha_se_necessario():
    if os.path.exists(ARQUIVO_EXCEL):
        return
    garantir_pastas()
    wb = openpyxl.Workbook()

    ws_v = wb.active; ws_v.title = "VENDAS"
    ws_v.append(CAB_VENDAS)
    _estilizar_cab_oxl(ws_v, 1, len(CAB_VENDAS), "365F92")
    ws_v.row_dimensions[1].height = 22
    _criar_tabela_oxl(ws_v, "TabelaVendas", f"A1:{get_column_letter(len(CAB_VENDAS))}2", "TableStyleMedium9")
    _resumo_mensal_oxl(ws_v)
    _ajustar_colunas(ws_v, {"A":6,"B":12,"C":22,"D":26,"E":12,"F":10,
                             "G":10,"H":14,"I":14,"J":14,"K":14,"L":14})

    ws_e = wb.create_sheet("ESTOQUE")
    ws_e.append(CAB_ESTOQUE)
    _estilizar_cab_oxl(ws_e, 1, len(CAB_ESTOQUE), "1B5E20")
    ws_e.row_dimensions[1].height = 22
    _criar_tabela_oxl(ws_e, "TabelaEstoque", f"A1:{get_column_letter(len(CAB_ESTOQUE))}2", "TableStyleMedium2")
    _ajustar_colunas(ws_e, {"A":12,"B":10,"C":28,"D":16,"E":12,"F":10,
                             "G":10,"H":14,"I":14,"J":38,"K":20})
    wb.save(ARQUIVO_EXCEL)


# ============================================================
# BACKUP & RESTAURAÇÃO
# ============================================================
def fazer_backup():
    if not os.path.exists(ARQUIVO_EXCEL):
        messagebox.showwarning("Aviso", "Nenhuma planilha encontrada."); return
    garantir_pastas()
    ts = datetime.now().strftime("%d%m%Y_%H%M%S")
    dest = os.path.join(PASTA_BACKUP, f"backup_{ts}.xlsx")
    try:
        shutil.copy2(ARQUIVO_EXCEL, dest)
        messagebox.showinfo("✅ Backup Salvo!", f"Backup criado!\n📁 {dest}")
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível fazer backup:\n{e}")

def restaurar_backup():
    garantir_pastas()
    arquivos = sorted([f for f in os.listdir(PASTA_BACKUP) if f.endswith(".xlsx")], reverse=True)
    if not arquivos:
        messagebox.showinfo("Backups", "Nenhum backup encontrado."); return

    win = tk.Toplevel(root)
    win.title("📂 Restaurar Backup"); win.geometry("500x380")
    win.configure(bg="#f5f5f5"); win.grab_set(); win.resizable(False, False)
    tk.Label(win, text="📂  RESTAURAR BACKUP", font=("Arial",13,"bold"),
             bg="#f5f5f5", fg="#1b5e20").pack(pady=12)
    tk.Label(win, text="Selecione o backup que deseja restaurar:",
             font=("Arial",10), bg="#f5f5f5").pack()
    fl = tk.Frame(win, bg="#f5f5f5"); fl.pack(fill="both", expand=True, padx=20, pady=10)
    lb = tk.Listbox(fl, font=("Consolas",10), height=10, selectmode="single",
                     relief="groove", bd=2, selectbackground="#1b5e20", selectforeground="white")
    sc = ttk.Scrollbar(fl, orient="vertical", command=lb.yview)
    lb.configure(yscrollcommand=sc.set)
    lb.pack(side="left", fill="both", expand=True); sc.pack(side="right", fill="y")
    for a in arquivos: lb.insert(tk.END, a)
    lb.selection_set(0)

    def confirmar():
        sel = lb.curselection()
        if not sel: messagebox.showwarning("Aviso","Selecione um backup."); return
        path_bk = os.path.join(PASTA_BACKUP, arquivos[sel[0]])
        if not messagebox.askyesno("⚠ Confirmar",
            f"Isso substituirá a planilha atual!\n\nBackup: {arquivos[sel[0]]}\n\nContinuar?"): return
        if os.path.exists(ARQUIVO_EXCEL):
            ts = datetime.now().strftime("%d%m%Y_%H%M%S")
            shutil.copy2(ARQUIVO_EXCEL, os.path.join(PASTA_BACKUP, f"pre_restore_{ts}.xlsx"))
        try:
            shutil.copy2(path_bk, ARQUIVO_EXCEL)
            win.destroy()
            messagebox.showinfo("✅ Restaurado!",
                                 "Planilha restaurada!\n(Backup da versão anterior foi salvo.)")
            migrar_se_necessario()
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível restaurar:\n{e}\n\nFeche o Excel e tente novamente.")

    tk.Button(win, text="✅  RESTAURAR SELECIONADO", command=confirmar,
              bg="#1b5e20", fg="white", font=("Arial",12,"bold"),
              padx=20, pady=8, cursor="hand2", bd=0).pack(pady=10)
    tk.Label(win, text=f"📁 {PASTA_BACKUP}",
             font=("Arial",8,"italic"), bg="#f5f5f5", fg="#777").pack(pady=(0,8))


# ============================================================
# FUNÇÕES DE ESTOQUE — LEITURA (openpyxl)
# ============================================================
def _linha_para_dict(row):
    """Converte tuple de linha do ESTOQUE em dict."""
    return {
        "cod":         normalizar_cod(row[I_E_COD]),
        "cod_pai":     normalizar_cod(row[I_E_PAI]) if row[I_E_PAI] else normalizar_cod(row[I_E_COD]),
        "nome":        str(row[I_E_NOME] or ""),
        "categoria":   str(row[I_E_CAT]  or ""),
        "cor":         str(row[I_E_COR]  or ""),
        "tamanho":     str(row[I_E_TAM]  or ""),
        "quantidade":  int(para_float(row[I_E_QTD])),
        "preco_custo": para_float(row[I_E_CUSTO]),
        "preco_venda": para_float(row[I_E_VENDA]),
        "foto":        str(row[I_E_FOTO] or ""),
    }

def _carregar_estoque():
    """Retorna lista de dicts com todos os produtos."""
    criar_planilha_se_necessario()
    produtos = []
    try:
        wb = openpyxl.load_workbook(ARQUIVO_EXCEL, read_only=True, data_only=True)
        ws = wb["ESTOQUE"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None: continue
            # suporte a planilhas com 8 colunas (migradas mas ainda abertas)
            row_ext = list(row) + [None] * (11 - len(row))
            produtos.append(_linha_para_dict(row_ext))
        wb.close()
    except Exception as e:
        print(f"Erro ao carregar estoque: {e}")
    return produtos

def buscar_produto_por_codigo(cod):
    """Busca variante exata pelo código."""
    cod_norm = normalizar_cod(cod)
    if not cod_norm: return None
    for p in _carregar_estoque():
        if p["cod"] == cod_norm:
            return p
    return None

def buscar_variantes_por_pai(cod_pai):
    """Retorna lista de variantes com o mesmo COD_PAI."""
    cod_norm = normalizar_cod(cod_pai)
    if not cod_norm: return []
    return [p for p in _carregar_estoque() if p["cod_pai"] == cod_norm]


# ============================================================
# FUNÇÕES DE ESTOQUE — ESCRITA (xlwings)
# ============================================================
def _linha_produto_xw(ws_xw, cod_norm):
    ultima = xw_ultima_linha(ws_xw)
    for i in range(2, ultima + 1):
        val = ws_xw.range(f"A{i}").value
        if val is not None and normalizar_cod(val) == cod_norm:
            return i
    return None

def reduzir_estoque_xw(cod_produto, quantidade=1):
    cod_norm = normalizar_cod(cod_produto)
    try:
        wb, fechar = xw_abrir()
        ws = wb.sheets["ESTOQUE"]
        linha = _linha_produto_xw(ws, cod_norm)
        if linha is None:
            xw_salvar_fechar(wb, fechar); return None, 0
        qtd_atual = int(para_float(ws.range(f"G{linha}").value))
        if qtd_atual < quantidade:
            xw_salvar_fechar(wb, fechar); return False, qtd_atual
        nova = qtd_atual - quantidade
        ws.range(f"G{linha}").value = nova
        ws.range(f"K{linha}").value = datetime.now().strftime("%d/%m/%Y %H:%M")
        xw_salvar_fechar(wb, fechar)
        return True, nova
    except Exception as e:
        print(f"Erro ao reduzir estoque: {e}")
        return False, 0


# ============================================================
# POPUP — SELETOR DE VARIANTE
# ============================================================
def popup_selecionar_variante(variantes, parent_win, callback):
    """
    Mostra popup com lista de variantes (cor/tamanho/qtd).
    Chama callback(variante_dict) ao confirmar.
    """
    win = tk.Toplevel(parent_win)
    win.title("🎨 Selecionar Variante")
    win.geometry("760x340")
    win.configure(bg="#f5f5f5")
    win.grab_set(); win.resizable(False, False)

    tk.Label(win, text="Selecione a Cor / Tamanho desejado:",
             font=("Arial",11,"bold"), bg="#f5f5f5", fg="#1b5e20").pack(pady=10)

    frame_t = tk.Frame(win, bg="#f5f5f5"); frame_t.pack(fill="both", expand=True, padx=16, pady=(0,8))
    cols = ("COD","NOME","COR","TAMANHO","ESTOQUE","P.VENDA")
    tree = ttk.Treeview(frame_t, columns=cols, show="headings", height=8)
    for col, w in zip(cols, [80, 180, 90, 80, 70, 90]):
        tree.heading(col, text=col); tree.column(col, width=w, anchor="center")
    sb = ttk.Scrollbar(frame_t, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=sb.set)
    tree.pack(side="left", fill="both", expand=True); sb.pack(side="right", fill="y")

    tree.tag_configure("sem_estoque", foreground="#aaa")

    for v in variantes:
        tag = "sem_estoque" if v["quantidade"] == 0 else ""
        tree.insert("", "end",
                     values=(v["cod"], v["nome"], v["cor"] or "-",
                             v["tamanho"] or "-", v["quantidade"],
                             f"R$ {v['preco_venda']:.2f}"),
                     tags=(tag,))
    if variantes: tree.selection_set(tree.get_children()[0])

    def confirmar(event=None):
        sel = tree.selection()
        if not sel: return
        idx = tree.index(sel[0])
        win.destroy()
        callback(variantes[idx])

    tree.bind("<Double-1>", confirmar)
    tk.Button(win, text="✅  SELECIONAR", command=confirmar,
              bg="#1b5e20", fg="white", font=("Arial",11,"bold"),
              padx=20, pady=7, cursor="hand2", bd=0).pack(pady=8)


# ============================================================
# JANELA DE ESTOQUE
# ============================================================
class JanelaEstoque:
    def __init__(self, parent):
        self.win = tk.Toplevel(parent)
        self.win.title("📦 Controle de Estoque — VJ FIT")
        self.win.geometry("1000x720")
        self.win.configure(bg='#f1f8f1')
        self.win.grab_set()
        self.foto_path_selecionada = None
        self._produto_editar = None

        fh = tk.Frame(self.win, bg="#1b5e20", height=55)
        fh.pack(fill="x"); fh.pack_propagate(False)
        tk.Label(fh, text="📦  CONTROLE DE ESTOQUE  —  VJ FIT  📦",
                 font=("Arial",14,"bold"), bg="#1b5e20", fg="white").pack(pady=14)

        style = ttk.Style()
        style.configure("TNotebook.Tab", font=("Arial",10,"bold"), padding=[12,6])
        self.nb = ttk.Notebook(self.win)
        self.nb.pack(fill="both", expand=True, padx=12, pady=8)

        self._aba_adicionar()
        self._aba_editar()
        self._aba_listar()
        self.atualizar_listagem()

    # ─────────────────────────────── ABA ADICIONAR
    def _aba_adicionar(self):
        frame = tk.Frame(self.nb, bg='#f1f8f1')
        self.nb.add(frame, text="  ➕ Adicionar Produto  ")

        col_e = tk.Frame(frame, bg='#f1f8f1')
        col_e.pack(side="left", fill="both", expand=True, padx=25, pady=15)
        col_d = tk.Frame(frame, bg='#f1f8f1')
        col_d.pack(side="right", padx=25, pady=15)

        # Layout em grid para melhor alinhamento
        campos = [
            ("Nº da Variante: *",   "entry_cod",   "ex: 1, 2, 3 ...  (único dentro do pai)"),
            ("Código Pai: *",       "entry_pai",   "ex: 1 = Short Nike, 2 = Camisa Nike"),
            ("Nome do Produto: *",  "entry_nome",  "ex: SHORT CANELADO FGA"),
            ("Categoria:",          "entry_cat",   "ex: SHORT"),
            ("Cor: *",              "entry_cor",   "ex: PRETO"),
            ("Tamanho: *",          "entry_tam",   "ex: M"),
            ("Quantidade: *",       "entry_qtd",   ""),
            ("Preço de Custo (R$):", "entry_custo", "ex: 45,00"),
            ("Preço de Venda (R$):", "entry_venda", "ex: 89,90"),
        ]
        for i, (lbl, attr, hint) in enumerate(campos):
            tk.Label(col_e, text=lbl, font=("Arial",10),
                     bg='#f1f8f1', anchor="w").grid(row=i*2,   column=0, sticky="w", pady=(5,0))
            e = tk.Entry(col_e, width=34, font=("Arial",11), bd=2, relief="groove")
            e.grid(row=i*2+1, column=0, sticky="we", ipady=3)
            if hint:
                tk.Label(col_e, text=hint, font=("Arial",8,"italic"),
                         bg='#f1f8f1', fg="#888").grid(row=i*2+1, column=1, padx=(6,0), sticky="w")
            setattr(self, attr, e)


        # Ao sair do campo Pai, sugere próximo nº de variante
        self.entry_pai.bind("<FocusOut>", self._sugerir_proximo_variante)
        self.entry_pai.bind("<Return>",   self._sugerir_proximo_variante)

        tk.Button(col_e, text="💾  SALVAR PRODUTO", command=self.salvar_produto,
                  bg="#2e7d32", fg="white", font=("Arial",12,"bold"),
                  padx=20, pady=10, cursor="hand2", bd=0).grid(
                      row=len(campos)*2+1, column=0, columnspan=2, sticky="we", pady=18)

        # Foto
        tk.Label(col_d, text="📷  Foto do Produto",
                 font=("Arial",11,"bold"), bg='#f1f8f1').pack(pady=(0,6))
        self.foto_prev = tk.Label(col_d, text="Clique em\n'Selecionar Foto'",
                                   bg="#e8f5e9", width=22, height=11,
                                   relief="groove", font=("Arial",10), fg="#555")
        self.foto_prev.pack()
        tk.Button(col_d, text="📂  Selecionar Foto", command=self.selecionar_foto,
                  bg="#1565c0", fg="white", font=("Arial",10,"bold"),
                  padx=8, pady=7, cursor="hand2", bd=0).pack(pady=(8,4), fill="x")
        tk.Button(col_d, text="🗑  Remover Foto", command=self.remover_foto,
                  bg="#b71c1c", fg="white", font=("Arial",10),
                  padx=8, pady=6, cursor="hand2", bd=0).pack(fill="x")
        tk.Label(col_d, text=(
            "💡 O Cód.Pai agrupa todas as\n"
            "variantes de um produto.\n\n"
            "Ex: Pai '1' = Short Nike\n"
            "  Var 1 = Preto M\n"
            "  Var 2 = Azul G\n"
            "\nPai '2' = Camisa Nike\n"
            "  Var 1 = Rosa P\n\n"
            "A foto é compartilhada\nentre variantes do mesmo Pai."
        ),
                 font=("Arial",8,"italic"), bg='#f1f8f1', fg="#888").pack(pady=(10,0))

    def _sugerir_proximo_variante(self, event=None):
        """Ao digitar o Código Pai, sugere automaticamente o próximo número de variante."""
        pai = self.entry_pai.get().strip()
        if not pai: return
        pai_norm = normalizar_cod(pai)
        variantes = buscar_variantes_por_pai(pai_norm)
        if not self.entry_cod.get().strip():
            proximo = len(variantes) + 1
            self.entry_cod.delete(0, tk.END)
            self.entry_cod.insert(0, str(proximo))

    def selecionar_foto(self):
        path = filedialog.askopenfilename(
            title="Selecionar foto",
            filetypes=[("Imagens","*.jpg *.jpeg *.png *.webp *.bmp"),("Todos","*.*")])
        if path:
            self.foto_path_selecionada = path
            self._preview(path, self.foto_prev)

    def remover_foto(self):
        self.foto_path_selecionada = None
        self.foto_prev.config(image='', text="Clique em\n'Selecionar Foto'", width=22, height=11)
        self.foto_prev.image = None

    def _preview(self, path, lbl, size=(185,185)):
        if not PIL_DISPONIVEL: lbl.config(text="✅ Foto\nselecionada", image=''); return
        try:
            img = Image.open(path); img.thumbnail(size)
            ph = ImageTk.PhotoImage(img)
            lbl.config(image=ph, text='', width=size[0], height=size[1]); lbl.image = ph
        except Exception as e:
            messagebox.showerror("Erro", f"Imagem inválida:\n{e}")

    def salvar_produto(self):
        cod    = self.entry_cod.get().strip()
        pai    = self.entry_pai.get().strip()
        nome   = self.entry_nome.get().strip().upper()
        cat    = self.entry_cat.get().strip().upper()
        cor    = self.entry_cor.get().strip().upper()
        tam    = self.entry_tam.get().strip().upper()
        qtd_s  = self.entry_qtd.get().strip()
        cst_s  = self.entry_custo.get().strip()
        vnd_s  = self.entry_venda.get().strip()

        if not cod or not pai or not nome or not qtd_s:
            messagebox.showerror("Erro","Código Variante, Código Pai, Nome e Quantidade são obrigatórios!"); return

        pai_norm = normalizar_cod(pai)
        var_norm = normalizar_cod(cod)   # número local da variante (1, 2, 3...)
        cod_norm = f"{pai_norm}-{var_norm}"  # chave interna: PAI-VARIANTE (ex: 1-1, 2-3)

        if buscar_produto_por_codigo(cod_norm):
            messagebox.showerror("Código Duplicado",
                                  f"Variante '{var_norm}' do Pai '{pai_norm}' já existe!\n"
                                  f"Use 'Editar Produto' para modificar."); return
        try:
            qtd = int(qtd_s)
            if qtd < 0: raise ValueError
        except: messagebox.showerror("Erro","Quantidade inválida."); return
        try:
            custo = float(cst_s.replace('.','').replace(',','.')) if cst_s else 0.0
            venda = float(vnd_s.replace('.','').replace(',','.')) if vnd_s else 0.0
        except: messagebox.showerror("Erro","Preço inválido. Exemplo: 49,90"); return

        # Foto: salva pelo COD_PAI para compartilhar entre variantes
        foto_dest = ""
        if self.foto_path_selecionada:
            garantir_pastas()
            ext = os.path.splitext(self.foto_path_selecionada)[1]
            foto_dest = os.path.join(PASTA_FOTOS, f"{pai_norm}{ext}")
            shutil.copy2(self.foto_path_selecionada, foto_dest)
        else:
            # Tenta herdar foto existente do pai
            vars_pai = buscar_variantes_por_pai(pai_norm)
            for v in vars_pai:
                if v["foto"] and os.path.exists(v["foto"]):
                    foto_dest = v["foto"]; break

        criar_planilha_se_necessario()
        try:
            wb, fechar = xw_abrir()
            ws = wb.sheets["ESTOQUE"]
            ul = xw_ultima_linha(ws) + 1
            ws.range(f"A{ul}").value = [
                cod_norm, pai_norm, nome, cat, cor, tam, qtd,
                custo, venda, foto_dest, datetime.now().strftime("%d/%m/%Y %H:%M")
            ]
            ws.range(f"H{ul}").number_format = FMT_MOEDA
            ws.range(f"I{ul}").number_format = FMT_MOEDA
            xw_salvar_fechar(wb, fechar)
            messagebox.showinfo("✅ Produto Adicionado!",
                                 f"'{nome}' — {cor} {tam}\n"
                                 f"Pai: {pai_norm}  |  Variante: {var_norm}  |  Código interno: {cod_norm}\n"
                                 f"Qtd: {qtd}  |  R$ {venda:.2f}")
            # Mantém o Código Pai para facilitar cadastro de múltiplas variantes
            for a in ["entry_cod","entry_nome","entry_cat",
                      "entry_cor","entry_tam","entry_qtd","entry_custo","entry_venda"]:
                getattr(self, a).delete(0, tk.END)
            # Sugere próximo número de variante automaticamente
            pai_atual = self.entry_pai.get().strip()
            if pai_atual:
                vars_exist = buscar_variantes_por_pai(normalizar_cod(pai_atual))
                self.entry_cod.insert(0, str(len(vars_exist) + 1))
            self.remover_foto()
            self.atualizar_listagem()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar:\n{e}")

    # ─────────────────────────────── ABA EDITAR
    def _aba_editar(self):
        frame = tk.Frame(self.nb, bg='#f1f8f1')
        self.nb.add(frame, text="  ✏️ Editar Produto  ")

        tk.Label(frame, text="Buscar pelo Código (variante ou pai):",
                 font=("Arial",11,"bold"), bg='#f1f8f1').pack(pady=(18,4))

        fb = tk.Frame(frame, bg='#f1f8f1'); fb.pack()
        self.entry_busca_ed = tk.Entry(fb, width=18, font=("Arial",13), bd=2, relief="groove")
        self.entry_busca_ed.pack(side="left", padx=6, ipady=4)
        self.entry_busca_ed.bind("<Return>", lambda e: self.buscar_editar())
        tk.Button(fb, text="🔍  Buscar", command=self.buscar_editar,
                  bg="#1565c0", fg="white", font=("Arial",11,"bold"),
                  padx=14, pady=6, cursor="hand2", bd=0).pack(side="left")

        # Info do produto encontrado
        self.lbl_info_ed = tk.Label(frame, text="Nenhum produto buscado ainda.",
                                     font=("Arial",10,"italic"), bg='#f1f8f1', fg="#777")
        self.lbl_info_ed.pack(pady=(6,2))

        # Campos editáveis
        sep = tk.Frame(frame, bg="#ddd", height=1); sep.pack(fill="x", padx=25, pady=6)
        fg = tk.Frame(frame, bg='#f1f8f1'); fg.pack(padx=30)

        campos_ed = [
            ("Nome:",           "ed_nome",  0, 0),
            ("Categoria:",      "ed_cat",   0, 2),
            ("Cor:",            "ed_cor",   1, 0),
            ("Tamanho:",        "ed_tam",   1, 2),
            ("Código Pai:",     "ed_pai",   2, 0),
            ("Preço de Venda:", "ed_venda", 2, 2),
            ("Preço de Custo:", "ed_custo", 3, 0),
        ]
        for lbl, attr, row, col in campos_ed:
            tk.Label(fg, text=lbl, font=("Arial",10), bg='#f1f8f1',
                     anchor="w").grid(row=row*2, column=col, sticky="w", padx=(0 if col==0 else 20, 0), pady=(6,0))
            e = tk.Entry(fg, width=22, font=("Arial",11), bd=2, relief="groove")
            e.grid(row=row*2+1, column=col, sticky="we", padx=(0 if col==0 else 20, 0), ipady=3)
            setattr(self, attr, e)

        # Quantidade — linha especial com +/- e set
        tk.Label(fg, text="Quantidade atual:", font=("Arial",10), bg='#f1f8f1').grid(
            row=8, column=0, sticky="w", pady=(10,0))
        self.lbl_qtd_atual = tk.Label(fg, text="—", font=("Arial",12,"bold"),
                                       bg='#f1f8f1', fg="#1b5e20")
        self.lbl_qtd_atual.grid(row=9, column=0, sticky="w")

        tk.Label(fg, text="Ajustar quantidade:", font=("Arial",10), bg='#f1f8f1').grid(
            row=8, column=2, sticky="w", padx=(20,0), pady=(10,0))
        fq2 = tk.Frame(fg, bg='#f1f8f1'); fq2.grid(row=9, column=2, sticky="w", padx=(20,0))
        tk.Label(fq2, text="Adicionar:", font=("Arial",9), bg='#f1f8f1').pack(side="left")
        self.ed_qtd_add = tk.Entry(fq2, width=6, font=("Arial",11), bd=2, relief="groove")
        self.ed_qtd_add.pack(side="left", padx=4, ipady=2)
        tk.Label(fq2, text="  Definir para:", font=("Arial",9), bg='#f1f8f1').pack(side="left")
        self.ed_qtd_set = tk.Entry(fq2, width=6, font=("Arial",11), bd=2, relief="groove")
        self.ed_qtd_set.pack(side="left", padx=4, ipady=2)

        # Botões de ação
        fb2 = tk.Frame(frame, bg='#f1f8f1'); fb2.pack(pady=14)
        tk.Button(fb2, text="💾  SALVAR ALTERAÇÕES", command=self.salvar_edicao,
                  bg="#1565c0", fg="white", font=("Arial",12,"bold"),
                  padx=22, pady=9, cursor="hand2", bd=0).pack(side="left", padx=8)
        tk.Button(fb2, text="🗑  DELETAR VARIANTE", command=self.deletar_variante_editar,
                  bg="#c62828", fg="white", font=("Arial",12,"bold"),
                  padx=22, pady=9, cursor="hand2", bd=0).pack(side="left", padx=8)

    def buscar_editar(self):
        cod = self.entry_busca_ed.get().strip()
        if not cod: messagebox.showerror("Erro","Digite um código."); return
        cod_norm = normalizar_cod(cod)

        # Tenta variante exata
        p = buscar_produto_por_codigo(cod_norm)
        if not p:
            # Tenta como código pai
            variantes = buscar_variantes_por_pai(cod_norm)
            if not variantes:
                self.lbl_info_ed.config(text=f"Código '{cod_norm}' não encontrado.", fg="#c62828")
                self._produto_editar = None; return
            if len(variantes) == 1:
                p = variantes[0]
            else:
                popup_selecionar_variante(variantes, self.win, self._preencher_edicao)
                return
        self._preencher_edicao(p)

    def _preencher_edicao(self, p):
        self._produto_editar = p
        self.lbl_info_ed.config(
            text=f"✅  Cod: {p['cod']}  |  Pai: {p['cod_pai']}  |  {p['nome']}  |  {p['cor']} {p['tamanho']}",
            fg="#1b5e20")
        self.lbl_qtd_atual.config(text=str(p['quantidade']))

        for attr, val in [("ed_nome", p['nome']), ("ed_cat", p['categoria']),
                           ("ed_cor",  p['cor']),  ("ed_tam", p['tamanho']),
                           ("ed_pai",  p['cod_pai']),
                           ("ed_venda", f"{p['preco_venda']:.2f}".replace('.',',')),
                           ("ed_custo", f"{p['preco_custo']:.2f}".replace('.',','))]:
            e = getattr(self, attr)
            e.delete(0, tk.END); e.insert(0, val)
        self.ed_qtd_add.delete(0, tk.END)
        self.ed_qtd_set.delete(0, tk.END)

    def salvar_edicao(self):
        if not self._produto_editar:
            messagebox.showerror("Erro","Busque um produto primeiro!"); return
        cod_norm = normalizar_cod(self._produto_editar['cod'])

        novo_nome  = self.ed_nome.get().strip().upper()
        novo_cat   = self.ed_cat.get().strip().upper()
        nova_cor   = self.ed_cor.get().strip().upper()
        novo_tam   = self.ed_tam.get().strip().upper()
        novo_pai   = normalizar_cod(self.ed_pai.get().strip())
        vnd_s      = self.ed_venda.get().strip()
        cst_s      = self.ed_custo.get().strip()
        qtd_add_s  = self.ed_qtd_add.get().strip()
        qtd_set_s  = self.ed_qtd_set.get().strip()

        if not novo_nome:
            messagebox.showerror("Erro","O nome não pode ficar vazio."); return

        try:
            novo_venda = float(vnd_s.replace('.','').replace(',','.')) if vnd_s else self._produto_editar['preco_venda']
        except: messagebox.showerror("Erro","Preço de venda inválido."); return

        try:
            novo_custo = float(cst_s.replace('.','').replace(',','.')) if cst_s else self._produto_editar['preco_custo']
        except: messagebox.showerror("Erro","Preço de custo inválido."); return

        try:
            wb, fechar = xw_abrir()
            ws = wb.sheets["ESTOQUE"]
            linha = _linha_produto_xw(ws, cod_norm)
            if linha is None:
                messagebox.showerror("Erro","Produto não encontrado na planilha."); xw_salvar_fechar(wb, fechar); return

            ws.range(f"B{linha}").value = novo_pai
            ws.range(f"C{linha}").value = novo_nome
            ws.range(f"D{linha}").value = novo_cat
            ws.range(f"E{linha}").value = nova_cor
            ws.range(f"F{linha}").value = novo_tam
            ws.range(f"H{linha}").value = novo_custo
            ws.range(f"H{linha}").number_format = FMT_MOEDA
            ws.range(f"I{linha}").value = novo_venda
            ws.range(f"I{linha}").number_format = FMT_MOEDA

            qtd_atual = int(para_float(ws.range(f"G{linha}").value))
            if qtd_set_s:
                try:
                    nova_qtd = int(qtd_set_s)
                    if nova_qtd < 0: raise ValueError
                    ws.range(f"G{linha}").value = nova_qtd
                except: messagebox.showerror("Erro","Qtd inválida."); xw_salvar_fechar(wb,fechar); return
            elif qtd_add_s:
                try:
                    add = int(qtd_add_s)
                    ws.range(f"G{linha}").value = qtd_atual + add
                except: messagebox.showerror("Erro","Qtd inválida."); xw_salvar_fechar(wb,fechar); return

            ws.range(f"K{linha}").value = datetime.now().strftime("%d/%m/%Y %H:%M")
            xw_salvar_fechar(wb, fechar)

            messagebox.showinfo("✅ Salvo!", f"'{novo_nome}' — {nova_cor} {novo_tam}\nAlterações salvas!")
            self.lbl_info_ed.config(text="Nenhum produto buscado ainda.", fg="#777")
            self.lbl_qtd_atual.config(text="—")
            self._produto_editar = None
            self.entry_busca_ed.delete(0, tk.END)
            self.atualizar_listagem()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar:\n{e}")

    def deletar_variante_editar(self):
        if not self._produto_editar:
            messagebox.showwarning("Aviso","Busque um produto primeiro."); return
        p = self._produto_editar
        if not messagebox.askyesno("⚠ Confirmar Exclusão",
            f"Deletar a variante abaixo?\n\n"
            f"Cód: {p['cod']}  —  {p['nome']}\n"
            f"Cor: {p['cor']}  |  Tam: {p['tamanho']}  |  Qtd: {p['quantidade']}\n\n"
            "Esta ação não pode ser desfeita!"): return
        cod_norm = normalizar_cod(p['cod'])
        try:
            wb, fechar = xw_abrir()
            ws = wb.sheets["ESTOQUE"]
            linha = _linha_produto_xw(ws, cod_norm)
            if linha is None:
                messagebox.showerror("Erro","Produto não encontrado."); xw_salvar_fechar(wb,fechar); return
            ws.range(f"{linha}:{linha}").delete()
            xw_salvar_fechar(wb, fechar)
            messagebox.showinfo("✅ Deletado!", f"Variante '{p['nome']} — {p['cor']} {p['tamanho']}' removida.")
            self.lbl_info_ed.config(text="Nenhum produto buscado ainda.", fg="#777")
            self.lbl_qtd_atual.config(text="—")
            self._produto_editar = None
            self.atualizar_listagem()
        except Exception as e:
            messagebox.showerror("Erro", "Erro ao deletar:\n" + str(e))

    # ─────────────────────────────── ABA LISTAR
    def _aba_listar(self):
        frame = tk.Frame(self.nb, bg='#f1f8f1')
        self.nb.add(frame, text="  📋 Listar Estoque  ")

        barra = tk.Frame(frame, bg='#f1f8f1'); barra.pack(fill="x", padx=10, pady=6)
        tk.Button(barra, text="🔄 Atualizar Lista", command=self.atualizar_listagem,
                  bg="#546e7a", fg="white", font=("Arial",10,"bold"),
                  padx=10, pady=5, cursor="hand2", bd=0).pack(side="left", padx=5)
        for cor, txt in [("#ffcdd2","⚠ Baixo (≤2)"),("#fff9c4","⚡ Médio (3-5)"),("#e8f5e9","✅ Ok (>5)")]:
            tk.Label(barra, text=txt, bg=cor, font=("Arial",9),
                     padx=6, pady=3, relief="groove").pack(side="left", padx=4)
        self.lbl_totais = tk.Label(barra, text="", font=("Arial",10,"bold"),
                                    bg='#f1f8f1', fg="#1b5e20")
        self.lbl_totais.pack(side="right", padx=10)

        # ── Barra de filtros ──────────────────────────────
        barra_filtros = tk.Frame(frame, bg='#e8f5e9', relief="groove", bd=1)
        barra_filtros.pack(fill="x", padx=10, pady=(0,6))

        def _campo_filtro(parent, label, width=12):
            fr = tk.Frame(parent, bg='#e8f5e9')
            fr.pack(side="left", padx=(8,0), pady=6)
            tk.Label(fr, text=label, font=("Arial",8,"bold"),
                     bg='#e8f5e9', fg="#1b5e20").pack(anchor="w")
            e = tk.Entry(fr, width=width, font=("Arial",10), bd=2, relief="groove")
            e.pack(ipady=2)
            return e

        tk.Label(barra_filtros, text="🔍 Filtros:", font=("Arial",9,"bold"),
                  bg='#e8f5e9', fg="#1b5e20").pack(side="left", padx=(8,0))
        self.filtro_nome = _campo_filtro(barra_filtros, "Nome", 16)
        self.filtro_cat  = _campo_filtro(barra_filtros, "Categoria", 12)
        self.filtro_cor  = _campo_filtro(barra_filtros, "Cor", 10)
        self.filtro_tam  = _campo_filtro(barra_filtros, "Tamanho", 8)
        self.filtro_preco_min = _campo_filtro(barra_filtros, "Preço mín.", 8)
        self.filtro_preco_max = _campo_filtro(barra_filtros, "Preço máx.", 8)

        for e in [self.filtro_nome, self.filtro_cat, self.filtro_cor,
                  self.filtro_tam, self.filtro_preco_min, self.filtro_preco_max]:
            e.bind("<KeyRelease>", lambda ev: self.aplicar_filtros())

        tk.Button(barra_filtros, text="✖ Limpar", command=self.limpar_filtros,
                  bg="#9e9e9e", fg="white", font=("Arial",9,"bold"),
                  padx=10, pady=5, cursor="hand2", bd=0).pack(side="left", padx=8, pady=6)

        cols = ("COD","PAI","NOME","COR","TAM","QTD","P.VENDA","FOTO")
        self.tree = ttk.Treeview(frame, columns=cols, show="headings", height=15)
        for col, w in zip(cols, [80,60,200,90,60,55,95,55]):
            self.tree.heading(col, text=col, command=lambda c=col: self._ordenar(c))
            self.tree.column(col, width=w, anchor="center")
        sb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=(10,0), pady=5)
        sb.pack(side="right", fill="y", pady=5)
        self.tree.tag_configure("baixo", background="#ffcdd2")
        self.tree.tag_configure("medio", background="#fff9c4")
        self.tree.tag_configure("ok",    background="#e8f5e9")
        self.tree.bind("<Double-1>", self.ver_foto)
        tk.Label(frame, text="💡 Duplo clique para ver a foto do produto",
                 font=("Arial",9,"italic"), bg='#f1f8f1', fg="#777").pack(pady=(0,4))

    def _ordenar(self, col):
        dados = [(self.tree.set(k,col),k) for k in self.tree.get_children('')]
        try: dados.sort(key=lambda t: float(t[0].replace('R$ ','').replace('.','').replace(',','.')))
        except: dados.sort()
        for i,(_,k) in enumerate(dados): self.tree.move(k,'',i)

    def atualizar_listagem(self):
        criar_planilha_se_necessario()
        try:
            self._produtos_cache = _carregar_estoque()
        except Exception as e:
            print(f"Erro ao listar: {e}")
            self._produtos_cache = []
        self.aplicar_filtros()

    def _renderizar_lista(self, produtos):
        for r in self.tree.get_children(): self.tree.delete(r)
        tp, tv = 0, 0.0
        for p in produtos:
            tf = "✅" if p["foto"] and os.path.exists(p["foto"]) else "❌"
            tp += p["quantidade"]; tv += p["quantidade"] * p["preco_venda"]
            tag = "baixo" if p["quantidade"]<=2 else ("medio" if p["quantidade"]<=5 else "ok")
            self.tree.insert("","end",
                              values=(p["cod"],p["cod_pai"],p["nome"],p["cor"],
                                      p["tamanho"],p["quantidade"],
                                      f"R$ {p['preco_venda']:.2f}",tf),
                              tags=(tag,))
        self.lbl_totais.config(text=f"Total: {tp} peças  |  Estoque: R$ {tv:.2f}")

    def aplicar_filtros(self):
        if not hasattr(self, "_produtos_cache"):
            self._produtos_cache = []
        nome = self.filtro_nome.get().strip().upper()
        cat  = self.filtro_cat.get().strip().upper()
        cor  = self.filtro_cor.get().strip().upper()
        tam  = self.filtro_tam.get().strip().upper()
        pmin_s = self.filtro_preco_min.get().strip()
        pmax_s = self.filtro_preco_max.get().strip()

        try: pmin = float(pmin_s.replace(',', '.')) if pmin_s else None
        except: pmin = None
        try: pmax = float(pmax_s.replace(',', '.')) if pmax_s else None
        except: pmax = None

        filtrados = []
        for p in self._produtos_cache:
            if nome and nome not in p["nome"].upper(): continue
            if cat  and cat  not in p["categoria"].upper(): continue
            if cor  and cor  not in p["cor"].upper(): continue
            if tam  and tam  not in p["tamanho"].upper(): continue
            if pmin is not None and p["preco_venda"] < pmin: continue
            if pmax is not None and p["preco_venda"] > pmax: continue
            filtrados.append(p)

        self._renderizar_lista(filtrados)

    def limpar_filtros(self):
        for e in [self.filtro_nome, self.filtro_cat, self.filtro_cor,
                  self.filtro_tam, self.filtro_preco_min, self.filtro_preco_max]:
            e.delete(0, tk.END)
        self.aplicar_filtros()

    def ver_foto(self, event):
        if not PIL_DISPONIVEL:
            messagebox.showinfo("Info","Instale Pillow:\npip install Pillow"); return
        item = self.tree.selection()
        if not item: return
        cod = str(self.tree.item(item[0])['values'][0])
        p = buscar_produto_por_codigo(cod)
        if not p or not p.get('foto') or not os.path.exists(p['foto']):
            messagebox.showinfo("Foto","Produto sem foto."); return
        wf = tk.Toplevel(self.win)
        wf.title(f"📷 {p['nome']}"); wf.configure(bg="white")
        wf.resizable(False,False); wf.grab_set()
        try:
            img = Image.open(p['foto']); img.thumbnail((520,520))
            ph = ImageTk.PhotoImage(img)
            tk.Label(wf, image=ph, bg="white").pack(padx=20, pady=12); wf._photo = ph
            tk.Label(wf, text=f"{p['nome']}  |  {p['cor']} {p['tamanho']}  |  "
                               f"Cód: {p['cod']}  |  Qtd: {p['quantidade']}  |  R$ {p['preco_venda']:.2f}",
                     font=("Arial",11,"bold"), bg="white", fg="#1b5e20").pack(pady=(0,12))
        except Exception as e:
            tk.Label(wf, text=f"Erro: {e}", bg="white").pack(padx=20, pady=20)


# ============================================================
# JANELA DE RESUMO
# ============================================================
MESES = {1:"Janeiro",2:"Fevereiro",3:"Março",4:"Abril",5:"Maio",6:"Junho",
         7:"Julho",8:"Agosto",9:"Setembro",10:"Outubro",11:"Novembro",12:"Dezembro"}

def abrir_resumo():
    win = tk.Toplevel(root)
    win.title("📊 Resumo de Vendas — VJ FIT")
    win.geometry("950x640"); win.configure(bg="#f5f5f5"); win.grab_set()

    fh2 = tk.Frame(win, bg="#1b5e20", height=52)
    fh2.pack(fill="x"); fh2.pack_propagate(False)
    tk.Label(fh2, text="📊  RESUMO DE VENDAS  —  VJ FIT",
             font=("Arial",14,"bold"), bg="#1b5e20", fg="white").pack(pady=14)

    ff = tk.Frame(win, bg="#f5f5f5"); ff.pack(fill="x", padx=20, pady=10)
    hoje = datetime.now()
    anos_disp = [str(a) for a in range(hoje.year-3, hoje.year+2)]
    meses_disp = [f"{v} ({k:02d})" for k,v in MESES.items()]

    tk.Label(ff, text="Mês:", font=("Arial",11,"bold"), bg="#f5f5f5").pack(side="left", padx=(0,5))
    combo_mes = ttk.Combobox(ff, values=meses_disp, width=18, font=("Arial",11), state="readonly")
    combo_mes.set(f"{MESES[hoje.month]} ({hoje.month:02d})")
    combo_mes.pack(side="left", padx=(0,15))
    tk.Label(ff, text="Ano:", font=("Arial",11,"bold"), bg="#f5f5f5").pack(side="left", padx=(0,5))
    combo_ano = ttk.Combobox(ff, values=anos_disp, width=8, font=("Arial",11), state="readonly")
    combo_ano.set(str(hoje.year))
    combo_ano.pack(side="left", padx=(0,15))
    tk.Button(ff, text="🔍  Filtrar", command=lambda: carregar_resumo(),
              bg="#1b5e20", fg="white", font=("Arial",11,"bold"),
              padx=14, pady=5, cursor="hand2", bd=0).pack(side="left")

    frame_cards = tk.Frame(win, bg="#f5f5f5"); frame_cards.pack(fill="x", padx=20, pady=(0,10))
    def make_card(titulo, cor):
        card = tk.Frame(frame_cards, bg=cor, padx=18, pady=10); card.pack(side="left", expand=True, fill="x", padx=6)
        tk.Label(card, text=titulo, font=("Arial",9,"bold"), bg=cor, fg="white").pack(anchor="w")
        lbl = tk.Label(card, text="—", font=("Arial",15,"bold"), bg=cor, fg="white"); lbl.pack(anchor="w")
        return lbl

    lbl_total  = make_card("💰 TOTAL VENDIDO",  "#1b5e20")
    lbl_vendas = make_card("🛒 Nº DE VENDAS",   "#1565c0")
    lbl_ticket = make_card("🎯 TICKET MÉDIO",   "#e65100")
    lbl_pecas  = make_card("📦 PEÇAS VENDIDAS", "#6a1b9a")

    tk.Frame(win, bg="#ddd", height=1).pack(fill="x", padx=20, pady=(0,6))

    frame_tree = tk.Frame(win, bg="#f5f5f5"); frame_tree.pack(fill="both", expand=True, padx=20, pady=(0,10))
    cols = ("ID","CLIENTE","PRODUTO","COR","TAM","QTD","V.UNIT","V.TOTAL","DATA","PAG","ENTREGA")
    tree = ttk.Treeview(frame_tree, columns=cols, show="headings", height=13)
    largs = {"ID":40,"CLIENTE":140,"PRODUTO":140,"COR":75,"TAM":55,
             "QTD":42,"V.UNIT":85,"V.TOTAL":85,"DATA":85,"PAG":80,"ENTREGA":85}
    for col in cols:
        tree.heading(col, text=col); tree.column(col, width=largs[col], anchor="center")
    sb = ttk.Scrollbar(frame_tree, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=sb.set)
    tree.pack(side="left", fill="both", expand=True); sb.pack(side="right", fill="y")
    tree.tag_configure("par",   background="#f9f9f9")
    tree.tag_configure("impar", background="#ffffff")

    def carregar_resumo():
        for r in tree.get_children(): tree.delete(r)
        mes_sel = combo_mes.get(); ano_sel = combo_ano.get()
        if not mes_sel or not ano_sel: return
        try:
            num_mes = int(mes_sel.split("(")[1].replace(")","").strip())
            num_ano = int(ano_sel)
        except: return

        total_valor = 0.0; total_vendas = 0; total_pecas = 0; idx = 0
        try:
            wb2 = openpyxl.load_workbook(ARQUIVO_EXCEL, read_only=True, data_only=True)
            ws2 = wb2["VENDAS"]
            for row in ws2.iter_rows(min_row=2, values_only=True):
                row_ext = list(row) + [None]*max(0, 12-len(row))
                data_val = row_ext[I_V_DATA]
                if data_val is None: continue
                if isinstance(data_val, datetime):
                    d = data_val
                elif isinstance(data_val, str):
                    try: d = datetime.strptime(data_val, "%d/%m/%Y")
                    except:
                        try: d = datetime.strptime(data_val, "%d/%m/%y")
                        except: continue
                else: continue
                if d.month != num_mes or d.year != num_ano: continue

                vid   = str(row_ext[0] or "")
                cli   = str(row_ext[2] or "")
                prod  = str(row_ext[3] or "")
                cor   = str(row_ext[4] or "")
                tam   = str(row_ext[5] or "")
                qtd   = int(para_float(row_ext[6]))
                vunit = para_float(row_ext[7])
                vtot  = para_float(row_ext[8])
                pag   = str(row_ext[10] or "")
                ent   = str(row_ext[11] or "")

                tag = "par" if idx%2==0 else "impar"
                tree.insert("","end",
                             values=(vid,cli,prod,cor,tam,qtd,
                                     f"R$ {vunit:.2f}",f"R$ {vtot:.2f}",
                                     d.strftime("%d/%m/%Y"),pag,ent),
                             tags=(tag,))
                total_valor += vtot; total_vendas += 1; total_pecas += qtd; idx += 1
            wb2.close()
        except Exception as e:
            print(f"Erro ao carregar resumo: {e}")

        ticket = total_valor/total_vendas if total_vendas > 0 else 0.0
        lbl_total.config( text=f"R$ {total_valor:.2f}")
        lbl_vendas.config(text=str(total_vendas))
        lbl_ticket.config(text=f"R$ {ticket:.2f}")
        lbl_pecas.config( text=str(total_pecas))

    carregar_resumo()
    combo_mes.bind("<<ComboboxSelected>>", lambda e: carregar_resumo())
    combo_ano.bind("<<ComboboxSelected>>", lambda e: carregar_resumo())


# ============================================================
# JANELA PRINCIPAL — VENDAS
# ============================================================
def abrir_estoque(): JanelaEstoque(root)

def _preencher_venda_com(p):
    """Preenche os campos de venda com os dados de uma variante."""
    entry_cod.delete(0,tk.END);     entry_cod.insert(0, p['cod'])
    entry_produto.delete(0,tk.END); entry_produto.insert(0, p['nome'])
    entry_cor.delete(0,tk.END);     entry_cor.insert(0, p['cor'])
    entry_tam.delete(0,tk.END);     entry_tam.insert(0, p['tamanho'])
    entry_valor.delete(0,tk.END);   entry_valor.insert(0, f"{p['preco_venda']:.2f}".replace('.',','))
    q = p['quantidade']
    cor = "#2e7d32" if q>5 else ("#e65100" if q>0 else "#c62828")
    lbl_info.config(text=f"✅  {p['nome']}  {p['cor']} {p['tamanho']}  |  Estoque: {q} un.", fg=cor)

def on_cod_change(event=None):
    cod = entry_cod.get().strip()
    if not cod: lbl_info.config(text="", fg="#555"); return
    cod_norm = normalizar_cod(cod)

    # VERIFICA PAI PRIMEIRO — garante popup mesmo quando COD == COD_PAI
    variantes = buscar_variantes_por_pai(cod_norm)

    if len(variantes) > 1:
        # Múltiplas variantes → popup para escolher cor/tamanho
        lbl_info.config(
            text=f"🎨  {len(variantes)} variante(s) encontrada(s) — selecione a cor/tamanho",
            fg="#1565c0")
        popup_selecionar_variante(variantes, root, _preencher_venda_com)

    elif len(variantes) == 1:
        # Apenas uma variante com esse pai → preenche direto
        _preencher_venda_com(variantes[0])

    else:
        # Sem filhos → tenta código de variante exato
        p = buscar_produto_por_codigo(cod_norm)
        if p:
            _preencher_venda_com(p)
        else:
            lbl_info.config(
                text=f"\u26a0  Código '{cod_norm}' não encontrado no estoque",
                fg="#c62828")

def salvar_pedido():
    if carrinho.vazio():
        messagebox.showwarning("Aviso", "Carrinho vazio")
        return

    nome = entry_nome.get().strip().upper()
    data_s = entry_data.get().strip()
    pag = combo_pag.get()
    ent = combo_ent.get()

    try:
        data_obj = datetime.strptime(data_s, "%d/%m/%y")
    except:
        messagebox.showerror("Erro", "Data inválida")
        return

    for item in carrinho.itens:
        p = buscar_produto_por_codigo(item.cod)
        if not p or p["quantidade"] < item.qtd:
            messagebox.showerror("Erro",
                                 f"{item.produto} sem estoque suficiente")
            return

    try:
        wb, fechar = xw_abrir()
        ws = wb.sheets["VENDAS"]

        ul = xw_ultima_linha(ws) + 1
        id_venda = ul - 1

        for item in carrinho.itens:
            ws.range(f"A{ul}").value = [
                id_venda,
                item.cod,
                nome,
                item.produto,
                item.cor,
                item.tam,
                item.qtd,
                item.valor,
                item.total,
                data_obj,
                pag,
                ent
            ]

            reduzir_estoque_xw(item.cod, item.qtd)
            ul += 1

        xw_salvar_fechar(wb, fechar)

        messagebox.showinfo("Sucesso",
                            f"Venda finalizada!\nTotal: R$ {carrinho.total():.2f}")

        carrinho.limpar()
        atualizar_carrinho_ui()
        limpar()

    except Exception as e:
        messagebox.showerror("Erro", str(e))

def atualizar_carrinho_ui():
    for i in tree_carrinho.get_children():
        tree_carrinho.delete(i)

    for idx, item in enumerate(carrinho.itens):
        tree_carrinho.insert("", "end", iid=str(idx), values=(
            item.produto,
            item.cor,
            item.tam,
            item.qtd,
            f"R$ {item.valor:.2f}",
            f"R$ {item.total:.2f}"
        ))

    total_itens = sum(i.qtd for i in carrinho.itens)

    lbl_total_carrinho.config(
        text=f"Itens: {total_itens} | Total: R$ {carrinho.total():.2f}"
    )


def adicionar_ao_carrinho():
    cod = entry_cod.get().strip()
    prod = entry_produto.get().strip()
    cor  = entry_cor.get().strip()
    tam  = entry_tam.get().strip()

    try:
        qtd = int(entry_qtd.get())
        valor = float(entry_valor.get().replace(',', '.'))
    except:
        messagebox.showerror("Erro", "Quantidade ou valor inválido")
        return

    p = buscar_produto_por_codigo(cod)
    if not p:
        messagebox.showerror("Erro", "Produto não encontrado")
        return

# soma quantidade já no carrinho
    qtd_no_carrinho = 0
    for i in carrinho.itens:
        if i.cod == cod:
            qtd_no_carrinho += i.qtd

    if p["quantidade"] < (qtd + qtd_no_carrinho):
        messagebox.showerror(
            "Estoque insuficiente",
            f"Disponível: {p['quantidade']} | Já no carrinho: {qtd_no_carrinho}"
        )
        return
    
    item = ItemCarrinho(cod, prod, cor, tam, qtd, valor)
    carrinho.adicionar(item)

    atualizar_carrinho_ui()
    lbl_info.config(text=f"🛒 {prod} adicionado ao carrinho!", fg="#2e7d32")


def remover_item_carrinho():
    sel = tree_carrinho.selection()
    if not sel:
        return
    index = int(sel[0])
    carrinho.remover(index)
    atualizar_carrinho_ui()

def limpar_carrinho():
    carrinho.limpar()
    atualizar_carrinho_ui()

def limpar():
    for e in [entry_cod, entry_nome, entry_produto, entry_cor, entry_tam, entry_valor, entry_data]:
        e.delete(0, tk.END)
    entry_qtd.delete(0, tk.END); entry_qtd.insert(0,"1")
    hoje = datetime.now()
    entry_data.insert(0, f"{hoje.day:02d}/{hoje.month:02d}/{str(hoje.year)[-2:]}")
    combo_pag.set(""); combo_ent.set("")
    lbl_info.config(text="", fg="#555")


# ============================================================
# INTERFACE PRINCIPAL
# ============================================================
root = tk.Tk()
def on_close():
    if not carrinho.vazio():
        if not messagebox.askyesno(
            "Carrinho em aberto",
            "Você tem itens no carrinho.\nDeseja sair mesmo?"
        ):
            return
    root.destroy()

root.protocol("WM_DELETE_WINDOW", on_close)
root.title("🛍 VJ FIT — Vendas & Estoque")
root.geometry("910x910"); root.configure(bg='#f0f0f0'); root.resizable(False, False)

# Header
fh = tk.Frame(root, bg="#2c3e50", height=65); fh.pack(fill="x"); fh.pack_propagate(False)
tk.Label(fh, text="🛍  VJ FIT — CONTROLE DE VENDAS",
         font=("Arial",14,"bold"), bg="#2c3e50", fg="white").pack(side="left", padx=20, pady=14)
fb = tk.Frame(fh, bg="#2c3e50"); fb.pack(side="right", padx=8, pady=8)
for txt, cmd, cor in [("📦  ESTOQUE",  abrir_estoque,    "#e67e22"),
                       ("📊  Resumo",   abrir_resumo,     "#1b5e20"),
                       ("💾  Backup",   fazer_backup,     "#1565c0"),
                       ("📂  Restaurar",restaurar_backup, "#6a1b9a")]:
    tk.Button(fb, text=txt, command=cmd, bg=cor, fg="white", font=("Arial",10,"bold"),
              padx=9, pady=7, cursor="hand2", bd=0).pack(side="left", padx=3)

# Corpo
fc = tk.Frame(root, bg='#f0f0f0'); fc.pack(fill="both", padx=30, pady=8)

tk.Label(fc, text="Código do Produto (variante ou pai):", font=("Arial",10), bg='#f0f0f0').pack(anchor="w")
entry_cod = tk.Entry(fc, width=22, font=("Arial",11), bd=2, relief="groove")
entry_cod.pack(anchor="w", ipady=3)
entry_cod.bind("<KeyRelease>", on_cod_change)

lbl_info = tk.Label(fc, text="", font=("Arial",9,"italic"), bg='#f0f0f0', fg="#555")
# ───────────── CARRINHO UI ─────────────
frame_carrinho = tk.Frame(fc, bg='#f0f0f0')
frame_carrinho.pack(fill="both", pady=10)

tk.Label(frame_carrinho, text="🛒 Carrinho",
         font=("Arial",11,"bold"), bg='#f0f0f0').pack(anchor="w")

cols = ("PRODUTO", "COR", "TAM", "QTD", "VALOR", "TOTAL")

tree_carrinho = ttk.Treeview(frame_carrinho, columns=cols, show="headings", height=6)

for col, w in zip(cols, [180,80,60,50,80,90]):
    tree_carrinho.heading(col, text=col)
    tree_carrinho.column(col, width=w, anchor="center")

tree_carrinho.pack(fill="x")

lbl_total_carrinho = tk.Label(frame_carrinho,
                             text="Total: R$ 0,00",
                             font=("Arial",12,"bold"),
                             fg="#1b5e20",
                             bg='#f0f0f0')
lbl_total_carrinho.pack(anchor="e", pady=5)

tk.Button(frame_carrinho, text="🗑 Remover Item",
          command=lambda: remover_item_carrinho(),
          bg="#c62828", fg="white").pack(side="left", padx=5)
lbl_info.pack(anchor="w", pady=(1,4))

tk.Button(frame_carrinho, text="🧹 Limpar Carrinho",
          command=limpar_carrinho,
          bg="#6a1b9a", fg="white").pack(side="left", padx=5)

tk.Label(fc, text="Nome do Cliente: *", font=("Arial",10), bg='#f0f0f0').pack(anchor="w")
entry_nome = tk.Entry(fc, width=48, font=("Arial",11), bd=2, relief="groove")
entry_nome.pack(fill="x", ipady=3)

tk.Label(fc, text="Produto: *", font=("Arial",10), bg='#f0f0f0').pack(anchor="w", pady=(6,0))
entry_produto = tk.Entry(fc, width=48, font=("Arial",11), bd=2, relief="groove")
entry_produto.pack(fill="x", ipady=3)

# Cor + Tamanho lado a lado
fct = tk.Frame(fc, bg='#f0f0f0'); fct.pack(fill="x", pady=(6,0))
fc2 = tk.Frame(fct, bg='#f0f0f0'); fc2.pack(side="left", padx=(0,10))
tk.Label(fc2, text="Cor:", font=("Arial",10), bg='#f0f0f0').pack(anchor="w")
entry_cor = tk.Entry(fc2, width=18, font=("Arial",11), bd=2, relief="groove")
entry_cor.pack(ipady=3)
ft2 = tk.Frame(fct, bg='#f0f0f0'); ft2.pack(side="left")
tk.Label(ft2, text="Tamanho:", font=("Arial",10), bg='#f0f0f0').pack(anchor="w")
entry_tam = tk.Entry(ft2, width=10, font=("Arial",11), bd=2, relief="groove")
entry_tam.pack(ipady=3)

# Quantidade + Valor
fqv = tk.Frame(fc, bg='#f0f0f0'); fqv.pack(fill="x", pady=(6,0))
fq = tk.Frame(fqv, bg='#f0f0f0'); fq.pack(side="left", padx=(0,10))
tk.Label(fq, text="Quantidade: *", font=("Arial",10), bg='#f0f0f0').pack(anchor="w")
entry_qtd = tk.Entry(fq, width=10, font=("Arial",11), bd=2, relief="groove")
entry_qtd.insert(0,"1"); entry_qtd.pack(ipady=3)
fv = tk.Frame(fqv, bg='#f0f0f0'); fv.pack(side="left")
tk.Label(fv, text="Valor Unitário (R$): *", font=("Arial",10), bg='#f0f0f0').pack(anchor="w")
entry_valor = tk.Entry(fv, width=18, font=("Arial",11), bd=2, relief="groove")
entry_valor.pack(ipady=3)

tk.Label(fc, text="Data (DD/MM/AA): *", font=("Arial",10), bg='#f0f0f0').pack(anchor="w", pady=(6,0))
entry_data = tk.Entry(fc, width=18, font=("Arial",11), bd=2, relief="groove")
hoje = datetime.now()
entry_data.insert(0, f"{hoje.day:02d}/{hoje.month:02d}/{str(hoje.year)[-2:]}")
entry_data.pack(anchor="w", ipady=3)

fcb = tk.Frame(fc, bg='#f0f0f0'); fcb.pack(fill="x", pady=(6,0))
fp = tk.Frame(fcb, bg='#f0f0f0'); fp.pack(side="left", fill="x", expand=True, padx=(0,10))
tk.Label(fp, text="Pagamento: *", font=("Arial",10), bg='#f0f0f0').pack(anchor="w")
combo_pag = ttk.Combobox(fp, font=("Arial",11), state="readonly")
combo_pag['values'] = ("Cartão","Pix","Dinheiro"); combo_pag.pack(fill="x")
fe2 = tk.Frame(fcb, bg='#f0f0f0'); fe2.pack(side="left", fill="x", expand=True)
tk.Label(fe2, text="Entrega: *", font=("Arial",10), bg='#f0f0f0').pack(anchor="w")
combo_ent = ttk.Combobox(fe2, font=("Arial",11), state="readonly")
combo_ent['values'] = ("ENTREGUE","ENVIADO","NO AGUARDO"); combo_ent.pack(fill="x")

tk.Button(fc, text="➕ Adicionar ao Carrinho",
          command=adicionar_ao_carrinho,
          bg="#1565c0", fg="white").pack(fill="x", pady=5)

tk.Button(fc, text="💾 Finalizar Venda",
          command=salvar_pedido,
          bg="#2e7d32", fg="white",
          font=("Arial",12,"bold")).pack(fill="x", pady=5)

tk.Label(root,
         text="💡 Digite o código pai para escolher cor/tamanho  |  O sistema salva direto no Excel",
         font=("Arial",8,"italic"), bg='#f0f0f0', fg="#2e7d32").pack(pady=(0,8))

# Inicializa
criar_planilha_se_necessario()
migrar_se_necessario()
root.mainloop()
